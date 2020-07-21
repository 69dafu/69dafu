#!/usr/bin/env python 
# encoding: utf-8 

"""
@version: v1.0
@author: cxa
@contact: 1598828268@qq.com
@site: https://www.cnblogs.com/c-x-a
@software: PyCharm
@file: r_w_excel.py
@time: 2020/7/6 22:03
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook


def runing(name, sheet_name):
    wb = load_workbook(name)
    sheet = wb[sheet_name]
    all_case=[]
    for i in range(2, sheet.max_row+1):
        case=[]
        for j in range(1,sheet.max_column-1):
            # print(sheet.cell(row=i+1,column=j+1).value)
            case.append(sheet.cell(row=i, column=j).value)
        # print(case)

        all_case.append(case)
    # print(all_case)
    return all_case
if __name__ == '__main__':
    allcase=runing("r_w.xlsx", "recharge")
    print("所有的测试数据",allcase)
