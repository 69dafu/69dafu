#!/usr/bin/env python 
# encoding: utf-8 

"""
@version: v1.0
@author: cxa
@contact: 1598828268@qq.com
@site: https://www.cnblogs.com/c-x-a
@software: PyCharm
@file: run.py
@time: 2020/7/6 22:47
"""
from lesson_2.r_w_excel import runing
from lesson_2.http_request import http_request
from openpyxl import load_workbook
Token=None
def run():
    global Token
    all_case=runing("r_w.xlsx","recharge")
    # print("获取到的数据是：",all_case)
    for i in range(len(all_case)):
        test_data=all_case[i]
        ip = "http://120.78.128.25:8766"
        response=http_request(ip+test_data[4],eval(test_data[5]),token=Token,method=test_data[3])
        if test_data[1]=="登录":
            Token="Bearer "+response["data"]["token_info"]["token"]
        print('最后的结果值', response)
        wb = load_workbook("r_w.xlsx")
        sheet = wb["recharge"]
        sheet.cell(row=test_data[0]+1, column=8).value =str(response)
        actual={"code":response["code"],"msg":response["msg"]}
        print("实际值",type(actual))
        print("期望值",type(eval(test_data[6])))
        expect=eval(test_data[6])
        if expect==actual:
            print("测试用例执行通过")
            sheet.cell(row=test_data[0]+1,column=9).value="PASS"
        else:
            print("测试用例执行不通过")
            sheet.cell(row=test_data[0] + 1, column=9).value = "FAIL"
        wb.save("r_w.xlsx")




run()



# log_data=all_case[0]
# ip="http://120.78.128.25:8766"
# log_response=http_request(ip+log_data[4],eval(log_data[5]),token=None,method=log_data[3])
#
# for i in range(1,len(all_case)):
#     test_data=all_case[i]
#     expected=eval(test_data[6])
#     token=log_response["data"]["token_info"]["token"]
#     response=http_request(ip+test_data[4],eval(test_data[5]),token="Bearer "+token,method=test_data[3])
#     print("最后的结果值",response)
# from openpyxl import load_workbook
# wb=load_workbook("r_w.xlsx")
# sheet=wb["recharge"]
# sheet.cell(row=2,column=8).value=str(response)

